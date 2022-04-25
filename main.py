import os.path
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from configparser import ConfigParser

config = ConfigParser()
config.read('config.ini')

def get_file_path(ignore_last_opened_file_path=False) -> str | None:
    """Проверяет наличие файла по пути из конфига или просит выбрать файл"""

    last_opened_file_path = config['DEFAULT']['last_opened_file_path']
    # если недоступен последний открываемый файл, просим юзера открыть новый файл и если он корректен,
    # перезаписываем путь в файл конфига
    if not os.path.exists(last_opened_file_path) or ignore_last_opened_file_path:
        if not ignore_last_opened_file_path:
            messagebox.showinfo(
                title='А карточек-то нету!',
                message='Всенепременнейше нужно выбрать .xlsx файл с данными для отображения карточек, иначе никак!',
            )
        file_path = filedialog.askopenfilename(
            title='Выберите файл',
            filetypes=[('Файлы excel', '*.xlsx')],
            initialdir='./files',
        )

        if not file_path:
            messagebox.showinfo(
                title='Ахтунг!',
                message='Файл не выбран, а это значит, что всё, всему конец. Выберите файл, пока не поздно...',
            )
            return
    else:
        file_path = last_opened_file_path

    return file_path


def ask_about_show_again_attribute() -> bool:
    return messagebox.askyesno(
        'Вот ведь незадача...',
        '''Нет карточек, которые следовало бы снова показывать.
Снять у всех карточек признак "не показывать снова"?''',
    )


class EnglishWordsLearner:
    cards_font = 'Arial 18'
    font = 'Arial 14'

    def __init__(self):
        # основные атрибуты данных
        self.file_name = ''
        self.work_book = None
        self.sheet = None
        self.cards_gen = None
        self.current_card = []

        # конфигурация окна и элементов окна
        self.window = tk.Tk()
        self.window.title('Файл не выбран')
        self.window.geometry('800x400')
        self.menu = tk.Menu()
        self.options = tk.Menu(tearoff=0)
        self.options.add_command(label='Выбрать файл', command=lambda: self.update_data(True))
        self.menu.add_cascade(label='Опции', menu=self.options)
        self.window.config(menu=self.menu)
        self.cards_frame = tk.Frame()
        self.cards_frame.pack(pady=30)
        self.buttons_frame = tk.Frame(height=80)
        self.buttons_frame.pack()
        self.btn_to_next_card = tk.Button(
            master=self.buttons_frame,
            text='Следующая',
            command=self.get_next_card,
            font=EnglishWordsLearner.font
        )
        self.btn_set_not_show_again_attribute = tk.Button(
            master=self.buttons_frame,
            text='Не показывать снова',
            command=self.set_not_show_again_attribute,
            font=EnglishWordsLearner.font
        )
        self.phrase_card = tk.Label(master=self.cards_frame, font=EnglishWordsLearner.cards_font, height=6)
        self.phrase_card.bind('<Motion>', lambda event: self.show_translate(event))
        self.phrase_card.pack()
        self.translate_card = tk.Label(master=self.cards_frame, font=EnglishWordsLearner.cards_font)
        self.translate_card.pack()

        self.update_data()

    def update_data(self, ignore_last_opened_file_path=False):
        """Получает данные из файла и записывает их в атрибуты"""
        file_name = get_file_path(ignore_last_opened_file_path)
        if file_name:
            self.file_name = file_name
            self.get_data()
            if self.cards_gen:
                self.get_next_card()
                self.buttons_frame.pack()
                self.btn_to_next_card.grid(column=1, row=1, padx='20')
                self.btn_set_not_show_again_attribute.grid(column=2, row=1)
        else:
            self.cards_is_out_message(False)

    def get_data(self):
        # во избежание утечет памяти, сбрасываем текущие данные перед открытием нового файла
        if self.work_book:
            self.work_book.close()
            self.work_book = None
            self.sheet = None
            self.cards_gen = None
            self.current_card = []

        wb = load_workbook(filename=self.file_name, keep_links=False)
        sheet = wb.active

        # проверка валидности данных из файла
        if sheet.max_column < 3:
            messagebox.showinfo(
                title='Что-то тут не так...',
                message='''В выбранном файле обнаружено меньшее количество столбцов, нежели необходимо для корректной 
работы. Ознакомьтесь с мануалом и отредактируйте файл под нужный формат.''',
            )
            self.phrase_card.config(text='Нет данных для отображения. Выберите новый файл.')
            return

        config.set('DEFAULT', 'last_opened_file_path', self.file_name)
        with open('config.ini', mode='w') as config_file:
            config.write(config_file)

        if list(list(sheet.iter_cols(min_col=3, max_col=3, values_only=True))[0]).count(1) == 0:
            answer = ask_about_show_again_attribute()
            if answer:
                for row in sheet.rows:
                    row[2].value = 1
                wb.save(filename=self.file_name)
            else:
                self.cards_is_out_message()
                return

        self.window.title(f'Карточки из файла "{self.file_name}"')
        self.work_book = wb
        self.sheet = sheet
        self.cards_gen = sheet.rows

    def get_next_card(self):
        """Получает из генератора карточек следующую карточку и проверяет,
                есть ли еще записи с признаком "показывать снова"""
        if list(self.sheet.iter_cols(min_col=3, max_col=3, values_only=True))[0].count(1) > 0:
            try:
                while True:
                    next_card = next(self.cards_gen)
                    if next_card[2].value:
                        self.current_card = next_card
                        self.phrase_card.config(text=next_card[0].value)
                        self.translate_card.config(text='')
                        break
            except StopIteration:
                answer = messagebox.askyesno(
                    'Похоже, карточки закончились',
                    'Показать текущие карточки по второму кругу?',
                )
                if answer:
                    self.get_data()
                    if self.cards_gen:
                        self.get_next_card()
                else:
                    self.cards_is_out_message()

        else:
            answer = ask_about_show_again_attribute()
            if answer:
                for row in self.sheet.rows:
                    row[2].value = 1
                self.work_book.save(filename=self.file_name)
                self.get_data()
                if self.cards_gen:
                    self.get_next_card()
            else:
                self.cards_is_out_message()

    def cards_is_out_message(self, show_message=True):
        if show_message:
            messagebox.showinfo(
                title='Ну, раз вы так!..',
                message='Что ж, тогда выберите другой файл с данными',
            )
        self.current_card = []
        self.phrase_card.config(text='Нет данных для отображения. Выберите новый файл.')
        self.translate_card.config(text='')
        self.buttons_frame.forget()

    def set_not_show_again_attribute(self):
        self.current_card[2].value = 0
        self.work_book.save(filename=self.file_name)
        self.get_next_card()

    def show_translate(self, args):
        if self.current_card:
            self.translate_card.config(text=self.current_card[1].value)

    def main(self):
        self.window.mainloop()


if __name__ == '__main__':
    learner = EnglishWordsLearner()
    learner.main()
